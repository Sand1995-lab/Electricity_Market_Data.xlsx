#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import requests
import pandas as pd
import datetime
import os
import logging
import time
import schedule
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)

def fetch_year_data(year):
    """Fetch data for a specific year from EIA website"""
    try:
        url = f"https://www.eia.gov/electricity/wholesalemarkets/csv/pjm_lmp_da_hr_zones_{year}.csv"
        logger.info(f"Fetching data from {url}")
        
        response = requests.get(url)
        response.raise_for_status()  # Raise exception for HTTP errors
        
        # Skip first 3 rows like in the original code
        df = pd.read_csv(url, skiprows=3)
        logger.info(f"{year} Data fetched successfully with {len(df)} rows")
        
        return df
    except Exception as e:
        logger.error(f"Error processing {year} data: {str(e)}")
        return pd.DataFrame()

def parse_date(date_time_str):
    """Parse date string into datetime object"""
    try:
        if not date_time_str or not isinstance(date_time_str, str):
            return None
        
        # Convert the date string to datetime object
        return pd.to_datetime(date_time_str)
    except Exception as e:
        logger.error(f"Error parsing date: {date_time_str}, Error: {str(e)}")
        return None

def create_combined_data(data_2024, data_2025, excel_file="Electricity_Market_Data.xlsx"):
    """Create combined data sheet with rolling year data"""
    try:
        # Get current date and one year ago date
        now = datetime.datetime.now()
        one_year_ago = now - datetime.timedelta(days=365)
        
        logger.info(f"Creating combined data for period: {one_year_ago.strftime('%Y-%m-%d')} to {now.strftime('%Y-%m-%d')}")
        
        # Check if we have valid data sets
        if data_2024.empty and data_2025.empty:
            logger.warning("No data available for either 2024 or 2025.")
            return
        
        # Determine which dataset to use for column names
        if not data_2025.empty:
            header_columns = data_2025.columns
        else:
            header_columns = data_2024.columns
        
        # Combine data from both years
        combined_df = pd.concat([data_2024, data_2025], ignore_index=True)
        
        # Ensure date column is in datetime format
        date_column = combined_df.columns[0]  # Assuming first column is the date column
        combined_df[date_column] = pd.to_datetime(combined_df[date_column], errors='coerce')
        
        # Filter for the rolling year period
        rolling_year_data = combined_df[(combined_df[date_column] >= one_year_ago) & 
                                        (combined_df[date_column] <= now)]
        
        # Sort data chronologically
        rolling_year_data = rolling_year_data.sort_values(by=date_column)
        
        # Calculate averages for numeric columns only
        numeric_columns = rolling_year_data.select_dtypes(include=['number']).columns
        averages = rolling_year_data[numeric_columns].mean().round(2)
        
        # Create a directory for the Excel file if it doesn't exist
        os.makedirs(os.path.dirname(os.path.abspath(excel_file)) if os.path.dirname(excel_file) else '.', exist_ok=True)
        
        # First, save each dataset to the Excel file
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
            if not data_2024.empty:
                data_2024.to_excel(writer, sheet_name="2024 Data", index=False)
                logger.info("2024 data written to Excel")
                
            if not data_2025.empty:
                data_2025.to_excel(writer, sheet_name="2025 Data", index=False)
                logger.info("2025 data written to Excel")
                
            # Write combined data
            rolling_year_data.to_excel(writer, sheet_name="Combined Data", index=False)
            logger.info(f"Combined data written to Excel with {len(rolling_year_data)} rows")
        
        # Now add the formatted averages row in a separate step
        # Load the workbook
        workbook = load_workbook(excel_file)
        if "Combined Data" in workbook.sheetnames:
            sheet = workbook["Combined Data"]
            row_num = len(rolling_year_data) + 2  # +2 for header and 0-indexing
            
            # Write "AVERAGE" in first column
            sheet.cell(row=row_num, column=1, value="AVERAGE")
            
            # Write average values for numeric columns
            for col_name, avg_value in averages.items():
                # Find the column index for this column in the original data
                if col_name in header_columns:
                    col_idx = list(header_columns).index(col_name) + 1  # +1 because Excel is 1-indexed
                    sheet.cell(row=row_num, column=col_idx, value=avg_value)
                    
                    # Format as bold with 2 decimal places
                    sheet.cell(row=row_num, column=col_idx).font = Font(bold=True)
                    sheet.cell(row=row_num, column=col_idx).number_format = "#,##0.00"
            
            # Format the entire row with light sky blue background
            light_sky_blue = "87CEFA"
            for col in range(1, len(header_columns) + 1):
                sheet.cell(row=row_num, column=col).fill = PatternFill(start_color=light_sky_blue, 
                                                                      end_color=light_sky_blue,
                                                                      fill_type="solid")
                sheet.cell(row=row_num, column=col).font = Font(bold=True)
            
            logger.info("Averages calculated and formatted with light sky blue background")
            
            # Save the workbook with the changes
            workbook.save(excel_file)
            logger.info(f"Excel file updated with averages: {excel_file}")
        
    except Exception as e:
        logger.error(f"Error creating combined data: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())

def update_data():
    """Main function to update all data"""
    try:
        logger.info("Starting complete data update process...")
        
        # Fetch data for both years
        data_2024 = fetch_year_data(2024)
        data_2025 = fetch_year_data(2025)
        
        # Create combined rolling year data
        create_combined_data(data_2024, data_2025)
        
        logger.info("Complete data update process finished successfully")
        return True
    except Exception as e:
        logger.error(f"Error in update process: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def schedule_updates():
    """Set up scheduled updates to run twice daily"""
    logger.info("Setting up scheduled updates to run at 5am and 5pm")
    
    # Schedule jobs to run at 5am and 5pm
    schedule.every().day.at("05:00").do(update_data)
    schedule.every().day.at("17:00").do(update_data)
    
    logger.info("Scheduled jobs set up successfully. Press Ctrl+C to exit.")
    
    # Run update immediately on startup
    update_data()
    
    # Keep the script running and check for scheduled jobs
    while True:
        schedule.run_pending()
        time.sleep(60)  # Check every minute

if __name__ == "__main__":
    # Run in scheduled mode or run once based on argument
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--once":
        # Run once mode
        success = update_data()
        if success:
            print("Data updated successfully. Excel file has been created/updated.")
        else:
            print("Failed to update data. See log for details.")
    else:
        # Run in scheduled mode
        schedule_updates()


# In[ ]:




